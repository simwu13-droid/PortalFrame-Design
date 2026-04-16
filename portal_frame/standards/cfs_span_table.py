"""CFS span table data loader and lookup (Formsteel CFS_Span_Table.xlsx).

Provides φN_c (compression) and φM_bx (major-axis bending) capacities as a
function of effective length, for Formsteel cold-formed steel sections.

The xlsx has two sheets:
  - "P (kN)"    : compression capacity, columns 1m..25m
  - "Mx (kN-m)" : bending capacity, columns 1m..25m

Library section names use a different convention from the span table, so a
static mapping LIBRARY_TO_SPANTABLE bridges them. Sections with no entry
return None (caller decides how to report).
"""

import os
import sys
import openpyxl


# ── Section name mapping: SpaceGass library name -> span table row name ──
#
# In the 63020 family, the span table convention is:
#   N         = nested base section
#   NS1 / NS2 = nested + 1 or 2 stiffener variants
LIBRARY_TO_SPANTABLE: dict[str, str] = {
    "270115":      "G550 270115",
    "270115N":     "G550 270115N",
    "290195":      "G550 290195",
    "290195N":     "G550 290195N",
    "50020":       "G550 50020",
    "50020N":      "G550 50020N",
    "63020N":      "G550 63020N",
    "63020S1":     "G550 63020NS1",
    "63020S2":     "G550 63020NS2",
    "440180195":   "Superspan 440x180x1.95",
    "440180195S2": "Superspan 440x180x1.95 2s",
    "650180295":   "Superspan 650x180x2.95",
    "650180295S1": "Superspan 650x180x2.95 1S",
    "650180295S2": "Superspan 650x180x2.95 2S",
}


_BUNDLE_DIR = getattr(sys, "_MEIPASS", None)
_PKG_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_DIR = os.path.abspath(os.path.join(_PKG_DIR, "..", ".."))

_SEARCH_PATHS = [
    os.path.join(_PROJECT_DIR, "docs", "CFS_Span_Table.xlsx"),
]
if _BUNDLE_DIR:
    _SEARCH_PATHS.insert(0, os.path.join(_BUNDLE_DIR, "docs", "CFS_Span_Table.xlsx"))


# Module-level caches (filled on first access)
_NC_TABLE: dict[str, list[float]] | None = None
_MBX_TABLE: dict[str, list[float]] | None = None
_TABLE_LENGTHS: list[float] = []


def _find_xlsx() -> str:
    for p in _SEARCH_PATHS:
        if os.path.isfile(p):
            return p
    raise FileNotFoundError(
        "CFS_Span_Table.xlsx not found. Searched: " + ", ".join(_SEARCH_PATHS)
    )


def _load_sheet(ws) -> tuple[list[float], dict[str, list[float]]]:
    """Parse one span-table sheet. Returns (lengths_m, {section_name: [values]})."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # Header is ('Section', '1m', '2m', ...). Strip the trailing 'm' and parse.
    lengths: list[float] = []
    for cell in header[1:]:
        if cell is None:
            break
        s = str(cell).strip().lower().rstrip("m").strip()
        try:
            lengths.append(float(s))
        except ValueError:
            break

    table: dict[str, list[float]] = {}
    n = len(lengths)
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        values = []
        for v in row[1 : 1 + n]:
            if v is None:
                values.append(0.0)
            else:
                values.append(float(v))
        if len(values) == n:
            table[name] = values
    return lengths, table


def _ensure_loaded() -> None:
    global _NC_TABLE, _MBX_TABLE, _TABLE_LENGTHS
    if _NC_TABLE is not None and _MBX_TABLE is not None:
        return

    path = _find_xlsx()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    if "P (kN)" not in wb.sheetnames or "Mx (kN-m)" not in wb.sheetnames:
        raise ValueError(
            f"CFS_Span_Table.xlsx must contain sheets 'P (kN)' and 'Mx (kN-m)'. "
            f"Found: {wb.sheetnames}"
        )

    nc_lengths, nc = _load_sheet(wb["P (kN)"])
    mbx_lengths, mbx = _load_sheet(wb["Mx (kN-m)"])

    if nc_lengths != mbx_lengths:
        raise ValueError(
            "P (kN) and Mx (kN-m) sheets have inconsistent length headers"
        )

    _NC_TABLE = nc
    _MBX_TABLE = mbx
    _TABLE_LENGTHS = nc_lengths


def _interp(lengths: list[float], values: list[float], L_m: float) -> float:
    """Linear interpolation with endpoint clamping."""
    if L_m <= lengths[0]:
        return values[0]
    if L_m >= lengths[-1]:
        return values[-1]
    for i in range(len(lengths) - 1):
        L1, L2 = lengths[i], lengths[i + 1]
        if L1 <= L_m <= L2:
            v1, v2 = values[i], values[i + 1]
            t = (L_m - L1) / (L2 - L1)
            return v1 + t * (v2 - v1)
    return values[-1]  # unreachable


def has_data(library_name: str) -> bool:
    """True if the section has both compression and bending data."""
    if library_name not in LIBRARY_TO_SPANTABLE:
        return False
    _ensure_loaded()
    key = LIBRARY_TO_SPANTABLE[library_name]
    return key in _NC_TABLE and key in _MBX_TABLE


def phi_Nc(library_name: str, L_m: float) -> float | None:
    """Compression capacity φN_c (kN) at effective length L (m).

    Linear interpolation between integer-meter columns. Clamps below 1m
    and above 25m. Returns None if section has no span table mapping.
    """
    if library_name not in LIBRARY_TO_SPANTABLE:
        return None
    _ensure_loaded()
    key = LIBRARY_TO_SPANTABLE[library_name]
    if key not in _NC_TABLE:
        return None
    return _interp(_TABLE_LENGTHS, _NC_TABLE[key], L_m)


def phi_Mbx(library_name: str, L_m: float) -> float | None:
    """Major-axis bending capacity φM_bx (kNm) at effective length L (m).

    Linear interpolation between integer-meter columns. Clamps below 1m
    and above 25m. Returns None if section has no span table mapping.
    """
    if library_name not in LIBRARY_TO_SPANTABLE:
        return None
    _ensure_loaded()
    key = LIBRARY_TO_SPANTABLE[library_name]
    if key not in _MBX_TABLE:
        return None
    return _interp(_TABLE_LENGTHS, _MBX_TABLE[key], L_m)
